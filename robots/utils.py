from datetime import timedelta
import io
from typing import Dict, List, Any

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from .constants import SHEET_HEADERS, HEADER_FORMAT_DATA, EMPTY_REPORT


def get_format_header() -> tuple[Font, Alignment]:
    """Определяеn формат заголовка"""
    return Font(
        bold=HEADER_FORMAT_DATA.get('bold', False),
        name=HEADER_FORMAT_DATA.get('font_name', 'Calibri'),
        size=HEADER_FORMAT_DATA.get('font_size', 11)
    ), Alignment(
        horizontal=HEADER_FORMAT_DATA.get('horizontal', 'center'),
        vertical=HEADER_FORMAT_DATA.get('vertical', 'center'),
        wrap_text=HEADER_FORMAT_DATA.get('wrap_text', False)
    )


def create_xls_analytics(
    data: Dict[str, List[Dict[str, Any]]],
) -> io.BytesIO:
    """Создает Excel-файл с полученными данными"""
    # Создаем Excel-файл в памяти
    output = io.BytesIO()
    file = Workbook()

    if not data:
        # Если данных нет, сохраняем пустую книгу с одним пустым листом
        worksheet = file.active
        worksheet.title = EMPTY_REPORT
    else:
        # Удаляем стандартный лист, созданный по умолчанию
        default_sheet = file.active
        file.remove(default_sheet)

        header_font, header_alignment = get_format_header()

        for model, versions in data.items():
            # Добавляем новый лист для модели
            worksheet = file.create_sheet(title=model)

            # Записываем заголовки столбцов
            for col_num, header in enumerate(SHEET_HEADERS, start=1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.alignment = header_alignment

            # Записываем данные поверсионно
            for row_num, version_data in enumerate(versions, start=2):
                worksheet.cell(row=row_num, column=1, value=model)
                worksheet.cell(
                    row=row_num, column=2, value=version_data.get('version')
                )
                worksheet.cell(
                    row=row_num, column=3, value=version_data.get('count')
                )

            # Устанавливаем ширину столбцов
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 15
            worksheet.column_dimensions['C'].width = 25

    # Сохраняем файл в BytesIO
    file.save(output)
    # Сбрасываем указатель потока на начало
    output.seek(0)

    return output


def get_text_current_date(text) -> str:
    """Генерирует текст с текущей датой."""
    return text.format(timezone.now().strftime("%d.%m.%Y"))


def get_created_date(days_ago: int = 0) -> str:
    """Генерирует текст: текущая дата - days_ago."""
    return (timezone.now() - timedelta(days=days_ago)).isoformat()


def get_serial(model: str, version: str) -> str:
    """Генерирует серийный номер."""
    return f'{model}-{version}'
