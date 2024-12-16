from http import HTTPStatus
from io import BytesIO

from django.db.models import Count
from openpyxl import load_workbook

from .fixtures import BaseRobotTest, CONTENT_TYPE_XML
from robots.constants import SHEET_HEADERS, EMPTY_REPORT
from robots.models import Robot


class GetReportViewTest(BaseRobotTest):
    def get_xls_file(self, response):
        return load_workbook(
            filename=BytesIO(
                b''.join(response.streaming_content)
            ),
            data_only=True
        )

    def test_status_code(self):
        """Проверка статуса запроса получения отчета"""
        self.assertEqual(
            self.client.get(self.url).status_code,
            HTTPStatus.OK
        )

    def test_content_type(self):
        """Проверка соответствия типа файла отчета"""
        self.assertEqual(
            self.client.get(self.url)['Content-Type'],
            CONTENT_TYPE_XML
        )

    def test_sheet_headers(self):
        """Проверка на соответствие заголовков в каждом листе"""
        xls_file = self.get_xls_file(self.client.get(self.url))
        for sheet in xls_file.worksheets:
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(
                headers,
                SHEET_HEADERS,
                f'Заголовки в листе "{sheet.title}" не совпадают. '
                f'Ожидалось: {SHEET_HEADERS}, Получено: {headers}'
            )

    def test_robots_view_sheet_data(self):
        """Проверка корректности записей в каждом листе отчета"""
        xls_file = self.get_xls_file(self.client.get(self.url))
        sheets_data = []
        for sheet in xls_file.worksheets:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row:
                    row_dict = dict(zip(['model', 'version', 'count'], row))
                    sheets_data.append(row_dict)

        last_week_robots = (
            Robot.recent_objects
            .values('model', 'version')
            .annotate(count=Count('id'))
            .order_by('model', 'version')
        )
        self.assertEqual(len(last_week_robots), len(sheets_data))
        for i in range(len(last_week_robots)):
            self.assertEqual(last_week_robots[i], sheets_data[i])

    def test_robots_view_empty_report(self):
        """Проверка создания корректного пустого отчета,
        когда нет роботов за последнюю неделю"""
        # Удаляем всех роботов, созданных за последнюю неделю
        Robot.recent_objects.all().delete()

        xls_file = self.get_xls_file(self.client.get(self.url))
        self.assertEqual(len(xls_file.worksheets), 1)
        self.assertEqual(xls_file.worksheets[0].title, EMPTY_REPORT)
